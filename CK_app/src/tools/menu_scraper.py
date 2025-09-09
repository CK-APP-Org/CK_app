import pandas as pd
import openpyxl
from datetime import datetime
import re

def scrape_menu(file_path, sheet_name="週菜單"):
    """
    Scrapes menu data from an Excel file containing cafeteria menu information,
    using the Monday date as an anchor and calculating other days based on column positions.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet containing menu data
    
    Returns:
        list: List of dictionaries containing date, dish name, and price information
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    menu_data = []
    
    # Find the Monday date in column A (formatted as 2025/2/24)
    monday_date = None
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if isinstance(cell_value, datetime):
            monday_date = cell_value
            print(f"DEBUG: Found Monday date: {monday_date}")
            break
        elif isinstance(cell_value, str):
            # Try to parse date strings like "2025/2/24"
            try:
                if "/" in cell_value:
                    parts = cell_value.split("/")
                    if len(parts) == 3:
                        year, month, day = map(int, parts)
                        monday_date = datetime(year, month, day)
                        print(f"DEBUG: Parsed Monday date from string: {monday_date}")
                        break
            except Exception as e:
                print(f"DEBUG: Failed to parse date string '{cell_value}': {str(e)}")
    
    if not monday_date:
        raise ValueError("Could not find Monday date in column A")
    
    # Define the columns for each day
    day_columns = {
        0: 'A',  # Monday
        1: 'G',  # Tuesday
        2: 'M',  # Wednesday
        3: 'S',  # Thursday
        4: 'Y'   # Friday
    }
    
    # Process each day's menu
    for day_offset, col_letter in day_columns.items():
        current_date = monday_date.replace(day=monday_date.day + day_offset)
        print(f"DEBUG: Processing menu for {current_date.strftime('%A')} ({current_date})")
        
        # Find "項次" for this day's column
        item_row = None
        col_index = ord(col_letter) - ord('A') + 1
        
        for row in range(1, sheet.max_row):
            if sheet.cell(row=row, column=col_index).value == "項次":
                item_row = row
                print(f"DEBUG: Found '項次' at {col_letter}{row}")
                break
        
        if not item_row:
            print(f"DEBUG: Couldn't find '項次' in column {col_letter}, skipping this day")
            continue
        
        # Process items for this day
        current_row = item_row + 1
        while current_row <= sheet.max_row:
            number_cell = sheet[f"{col_letter}{current_row}"]
            
            # Check if cell contains a number
            if isinstance(number_cell.value, (int, float)):
                # Get dish name (cell to the right)
                dish_col = chr(ord(col_letter) + 1)
                dish_cell = sheet[f"{dish_col}{current_row}"]
                
                # If dish cell is blank, we've reached the end of this day's menu
                if not dish_cell.value:
                    break
                    
                # Get price (cell two to the right)
                price_col = chr(ord(col_letter) + 2)
                price_cell = sheet[f"{price_col}{current_row}"]
                
                # Add to menu data
                menu_item = {
                    'date': current_date,
                    'day': current_date.strftime('%A'),
                    'dish_number': number_cell.value,
                    'dish_name': dish_cell.value,
                    'price': price_cell.value if price_cell.value else None
                }
                print(f"DEBUG: Adding menu item: {menu_item}")
                menu_data.append(menu_item)
            
            current_row += 1
    
    print(f"DEBUG: Total menu items collected: {len(menu_data)}")
    return menu_data

# Example usage
if __name__ == "__main__":
    try:
        menu_data = scrape_menu("menu.xlsx")
        for item in menu_data:
            print(f"Date: {item['date']}, {item['dish_number']}. {item['dish_name']} - ${item['price']}")
    except Exception as e:
        print(f"Error processing menu: {str(e)}")
